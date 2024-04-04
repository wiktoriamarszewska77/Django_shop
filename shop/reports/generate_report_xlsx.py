from openpyxl import Workbook
from openpyxl.styles import Alignment
from order.models import Order
from io import BytesIO


def generate_xlsx_report(orders, data_parameters):
    wb = Workbook()
    ws = wb.active

    headers_data = {
        "order.id": ("Order ID", lambda order: order.id),
        "order.buyer": ("Order Buyer", lambda order: order.buyer.username),
        "order.street": ("Order Street", lambda order: order.street),
        "order.city": ("Order City", lambda order: order.city),
        "order.postcode": ("Order Postcode", lambda order: order.postcode),
        "order.date": ("Order Date", lambda order: order.date.replace(tzinfo=None)),
        "order.delivery.name": (
            "Order Delivery Name",
            lambda order: order.delivery.name,
        ),
        "order.delivery.price": (
            "Order Delivery Price",
            lambda order: order.delivery.price,
        ),
        "order.status": ("Order Status", lambda order: order.status),
    }

    for col, (header_key, (header, _)) in enumerate(headers_data.items(), start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for row, order in enumerate(orders, start=2):
        for col, (header_key, (_, get_value)) in enumerate(
            headers_data.items(), start=1
        ):
            if header_key in data_parameters:
                value = get_value(order)
                ws.cell(row=row, column=col, value=value)

    buffer = BytesIO()
    wb.save(buffer)

    buffer.seek(0)
    binary_data = buffer.read()

    buffer.close()

    return binary_data


def filter_orders_by_date(start_date, end_date, user_id):
    orders = Order.objects.filter(date__range=(start_date, end_date), buyer=user_id)
    return orders
