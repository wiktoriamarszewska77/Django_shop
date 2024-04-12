from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from reportlab.pdfgen import canvas
from order.models import Order


def generate_pdf_report(orders, data_parameters):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)

    param_mapping = {
        "order.id": "Order ID",
        "order.buyer": "Order Buyer",
        "order.street": "Order Street",
        "order.city": "Order City",
        "order.postcode": "Order Postcode",
        "order.date": "Order Date",
        "order.delivery.name": "Order Delivery Name",
        "order.delivery.price": "Order Delivery Price",
        "order.status": "Order Status",
    }

    y_position = 800
    page_counter = 0

    for order in orders:
        for param in data_parameters:
            if param in param_mapping:
                if param == "order.delivery.name":
                    pdf.drawString(
                        100,
                        y_position,
                        f"{param_mapping[param]}: {order.delivery.name}",
                    )
                elif param == "order.delivery.price":
                    pdf.drawString(
                        100,
                        y_position,
                        f"{param_mapping[param]}: {order.delivery.price}",
                    )
                else:
                    pdf.drawString(
                        100,
                        y_position,
                        f"{param_mapping[param]}: {getattr(order, param.split('.')[1])}",
                    )
                y_position -= 20

        y_position -= 20

        if (page_counter + 1) % 4 == 0:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y_position = 800
        page_counter += 1

    pdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()

    return pdf_data


def generate_xlsx_report(orders, data_parameters):
    wb = Workbook()
    ws = wb.active

    headers_data = {
        "order.id": ("Order ID", lambda order: order.id),
        "order.buyer": ("Order Buyer", lambda order: order.buyer.username),
        "order.street": ("Order Street", lambda order: order.street),
        "order.city": ("Order City", lambda order: order.city),
        "order.postcode": ("Order Postcode", lambda order: order.postcode),
        "order.date": ("Order Date", lambda order: order.date.replace(tzinfo=None)),
        "order.delivery.name": (
            "Order Delivery Name",
            lambda order: order.delivery.name,
        ),
        "order.delivery.price": (
            "Order Delivery Price",
            lambda order: order.delivery.price,
        ),
        "order.status": ("Order Status", lambda order: order.status),
    }

    for col, (header_key, (header, _)) in enumerate(headers_data.items(), start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for row, order in enumerate(orders, start=2):
        for col, (header_key, (_, get_value)) in enumerate(
            headers_data.items(), start=1
        ):
            if header_key in data_parameters:
                value = get_value(order)
                ws.cell(row=row, column=col, value=value)

    buffer = BytesIO()
    wb.save(buffer)

    buffer.seek(0)
    binary_data = buffer.read()

    buffer.close()

    return binary_data


def filter_orders_by_date(start_date, end_date, user_id):
    orders = Order.objects.filter(date__range=(start_date, end_date), buyer=user_id)
    return orders
